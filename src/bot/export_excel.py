import os
from datetime import datetime

import aiosqlite
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from bot.config import DB_PATH, EXPORT_DIR


async def export_to_excel() -> str:
    """خروجی اکسل از تمام آگهی‌ها"""
    os.makedirs(EXPORT_DIR, exist_ok=True)

    filename = os.path.join(
        EXPORT_DIR, f"divar_ads_{datetime.now():%Y-%m-%d_%H-%M-%S}.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Divar Ads"

    # هدر
    headers = ["Token", "Title", "URL", "District", "Text", "Special", "Created At"]
    ws.append(headers)

    # استایل هدر
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E40AF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # داده‌ها
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("""
            SELECT token, title, url, district, text, special, created_at
            FROM ads
            ORDER BY created_at DESC
        """) as cursor:
            rows = await cursor.fetchall()
            for row in rows:
                ws.append(
                    [
                        row["token"],
                        row["title"],
                        row["url"],
                        row["district"],
                        row["text"],
                        "بله" if row["special"] else "خیر",
                        row["created_at"],
                    ]
                )

    # عرض ستون‌ها
    widths = {"A": 28, "B": 45, "C": 55, "D": 20, "E": 80, "F": 10, "G": 22}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(filename)
    return filename


async def export_links_txt() -> str:
    """خروجی فایل TXT فقط لینک‌ها"""
    os.makedirs(EXPORT_DIR, exist_ok=True)

    filename = os.path.join(
        EXPORT_DIR, f"divar_links_{datetime.now():%Y-%m-%d_%H-%M-%S}.txt"
    )

    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute("""
            SELECT url FROM ads
            WHERE url IS NOT NULL AND url != ''
            ORDER BY created_at DESC
        """) as cursor:
            rows = await cursor.fetchall()

    with open(filename, "w", encoding="utf-8") as f:
        f.write(f"# Divar Ads Links - {datetime.now():%Y-%m-%d %H:%M:%S}\n")
        f.write(f"# Total: {len(rows)} links\n\n")
        for row in rows:
            f.write(row[0] + "\n")

    return filename
